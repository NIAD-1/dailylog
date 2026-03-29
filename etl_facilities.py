#!/usr/bin/env python3
"""
PMS Lagos — Unified Facility Database ETL Pipeline
===================================================
Extracts data from 15 Excel spreadsheets, normalizes facility names,
deduplicates records, and produces structured JSON files for Firestore import.

Output files:
  - etl_output/master_facilities.json
  - etl_output/inspections.json
  - etl_output/sanctions.json
  - etl_output/complaints.json
  - etl_output/documents.json
  - etl_output/file_registry.json
  - etl_output/dedup_report.json
"""

import os
import re
import json
import hashlib
from datetime import datetime
from collections import defaultdict

import openpyxl

# ─── Configuration ───────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "etl_output")

EXCEL_FILES = {
    "routine_surveillance": "PMS-LAGOS ROUNTINE SURVILLANCE(2016-DATE) (1) (1).xlsx",
    "gdp_inspected": "GDP UPDATED INSPECTED FACILITIES current one DIrector.xlsx",
    "gdp_uninspected": "CURRENT UNINSPECTED GDP UPDATE.xlsx",
    "glsi_central": "GLSI for Lagos Central.xlsx",
    "glsi_east": "GLSI for Lagos East.xlsx",
    "glsi_west": "GLSI for Lagos West.xlsx",
    "glsi_not_located": "GLSI NOT LOCATED.xlsx",
    "glsi_defaulters": "GLSI Defaulters.xlsx",
    "admin_fees": "ADMIN FEES ISSUED 2021-2025.xlsx",
    "revenue": "2025 REVENUE SHEET.xlsx",
    "defaulters": "DEFAULTERS OF PAYMENT ADMIN FINE 2021-2025.xlsx",
    "complaints": "CONSUMER COMPLAINTS.xlsx",
    "documents": "DOCUMENTS RECEIVED AND DISPATCHED IN 2025.xlsx",
    "office_files": "PMS LAGOS OFFICE FILE.xlsx",
}

# LGA → Zone mapping for Lagos
LGA_ZONE_MAP = {
    # Lagos Central
    "LAGOS ISLAND": "Lagos Central", "LAGOS MAINLAND": "Lagos Central",
    "ETI-OSA": "Lagos Central", "SURULERE": "Lagos Central", "APAPA": "Lagos Central",
    # Lagos East
    "KOSOFE": "Lagos East", "IKORODU": "Lagos East", "EPE": "Lagos East",
    "IBEJU-LEKKI": "Lagos East", "SHOMOLU": "Lagos East",
    # Lagos West
    "AGEGE": "Lagos West", "IFAKO-IJAIYE": "Lagos West", "IFAKO-IJAYE": "Lagos West",
    "ALIMOSHO": "Lagos West", "BADAGRY": "Lagos West", "OJO": "Lagos West",
    "AJEROMI-IFELODUN": "Lagos West", "AMUWO-ODOFIN": "Lagos West",
    "OSHODI-ISOLO": "Lagos West", "IKEJA": "Lagos West", "MUSHIN": "Lagos West",
}


# ─── Utility Functions ──────────────────────────────────────────────────────

def normalize_name(name):
    """Normalize a facility name for deduplication."""
    if not name:
        return ""
    name = str(name).upper().strip()
    # Remove trailing punctuation and common noise
    name = re.sub(r'[.,;:!]+$', '', name)
    # Collapse multiple spaces
    name = re.sub(r'\s+', ' ', name)
    # Remove common suffixes for matching
    name = name.replace('\n', ' ').replace('\r', ' ')
    return name.strip()


def normalize_address(addr):
    """Normalize an address string."""
    if not addr:
        return ""
    addr = str(addr).upper().strip()
    addr = re.sub(r'\s+', ' ', addr)
    addr = addr.replace('\n', ' ').replace('\r', ' ')
    return addr.strip()


def parse_date(val):
    """Attempt to parse various date formats → ISO string."""
    if not val:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Try common formats
    formats = [
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y",
        "%dTH %B, %Y", "%dTH %B %Y", "%dST %B, %Y", "%dST %B %Y",
        "%dND %B, %Y", "%dND %B %Y", "%dRD %B, %Y", "%dRD %B %Y",
        "%dth %B, %Y", "%dth %B %Y", "%dst %B, %Y", "%dst %B %Y",
        "%dnd %B, %Y", "%dnd %B %Y", "%drd %B, %Y", "%drd %B %Y",
        "%B %Y",
    ]
    # Remove ordinal suffixes for parsing
    cleaned = re.sub(r'(\d+)(ST|ND|RD|TH)', r'\1', s, flags=re.IGNORECASE)
    cleaned = cleaned.replace(',', '').strip()
    # Also remove day-of-week names if present
    cleaned = re.sub(r'^(MONDAY|TUESDAY|WEDNESDAY|THURSDAY|FRIDAY|SATURDAY|SUNDAY)\s+', '', cleaned, flags=re.IGNORECASE)

    for fmt in formats:
        try:
            dt = datetime.strptime(cleaned, fmt.replace('TH', '').replace('ST', '').replace('ND', '').replace('RD', ''))
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

    # Last resort: try to find any date-like pattern
    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"

    # Fallback: return as-is (the raw string) for manual review
    return s


def parse_amount(val):
    """Parse Nigerian naira amounts like '#200,000' or '₦1,500,000'."""
    if not val:
        return 0
    s = str(val).strip()
    s = s.replace('#', '').replace('₦', '').replace('N', '').replace(',', '').replace(' ', '')
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def cell_str(val):
    """Convert cell value to clean string."""
    if val is None:
        return ""
    s = str(val).strip()
    if s in ('None', 'NIL', 'N/A', 'nil', 'n/a', 'NA', 'na', '-'):
        return ""
    return s.replace('\n', ' ').replace('\r', ' ')


def make_id(parts):
    """Generate a stable ID from a list of parts."""
    key = "|".join(str(p) for p in parts)
    return hashlib.md5(key.encode()).hexdigest()[:12]


def read_workbook(key):
    """Load an Excel workbook by key."""
    filename = EXCEL_FILES[key]
    path = os.path.join(BASE_DIR, filename)
    if not os.path.exists(path):
        print(f"  ⚠  File not found: {filename}")
        return None
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


# ─── Master Facility Registry ───────────────────────────────────────────────

class FacilityRegistry:
    """Central registry for deduplicating and managing facility records."""

    def __init__(self):
        self.facilities = {}          # normalized_name → facility dict
        self.name_map = {}            # normalized_name → canonical name
        self.merge_log = []           # log of merge decisions

    def _match_key(self, name):
        """Generate a matching key for fuzzy dedup."""
        key = normalize_name(name)
        # Remove common business suffixes and noise
        noise = [
            ' LIMITED', ' LTD', ' PLC', ' NIG', ' NIGERIA',
            ' ENTERPRISES', ' ENTERPRISE', ' ENT', ' COMPANY', ' CO',
            ' INTERNATIONAL', ' INTL', ' INCORPORATED', ' INC',
            ' PHARMACY', ' PHARMACEUTICAL', ' PHARMACEUTICALS', ' PHARM',
            ' VENTURES', ' VENTURE', ' GLOBAL', ' SERVICES', ' SERVICE',
            ' STORES', ' STORE', ' SUPERMARKET', ' SUPERSTORES'
        ]
        # Sort by length descending to replace longer ones first
        for suffix in sorted(noise, key=len, reverse=True):
            key = key.replace(suffix, '')
        
        # Remove non-alphanumeric for very strict core matching
        key = re.sub(r'[^A-Z0-9]', '', key)
        return key.strip()

    def register(self, name, address="", contact="", phone="", email="",
                 activity_type="", file_number="", zone="", lga="",
                 facility_type="", contact_person=""):
        """Register a facility — deduplicates against existing records."""
        norm = normalize_name(name)
        if not norm:
            return None

        match_key = self._match_key(name)
        addr_norm = normalize_address(address)

        # Check for existing match
        existing_key = None
        for existing_norm, facility in self.facilities.items():
            # 1. Exact Name Match
            if existing_norm == norm:
                existing_key = existing_norm
                break
            
            # 2. Match Key (Name without suffixes)
            existing_match = self._match_key(facility["name"])
            if existing_match == match_key and match_key:
                existing_key = existing_norm
                self.merge_log.append({
                    "action": "merged_by_name",
                    "incoming": name,
                    "matched_to": facility["name"],
                    "match_key": match_key
                })
                break
                
            # 3. Address Match (if address is substantial)
            if addr_norm and len(addr_norm) > 15 and facility["address"] == addr_norm:
                existing_key = existing_norm
                self.merge_log.append({
                    "action": "merged_by_address",
                    "incoming": name,
                    "matched_to": facility["name"],
                    "address": addr_norm
                })
                break

        addr_norm = normalize_address(address)

        if existing_key:
            # Update existing record with any new data
            f = self.facilities[existing_key]
            if addr_norm and not f["address"]:
                f["address"] = addr_norm
            if contact_person and not f.get("contactPerson"):
                f["contactPerson"] = contact_person
            if phone and not f.get("phone"):
                f["phone"] = phone
            if email and not f.get("email"):
                f["email"] = email
            if file_number and not f.get("fileNumber"):
                f["fileNumber"] = file_number
            if zone and not f.get("zone"):
                f["zone"] = zone
            if lga and not f.get("lga"):
                f["lga"] = lga
            if activity_type and activity_type not in f.get("activityTypes", []):
                f.setdefault("activityTypes", []).append(activity_type)
            # Track aliases
            if norm != existing_key and name not in f.get("aliases", []):
                f.setdefault("aliases", []).append(name)
            return f["id"]
        else:
            # Create new record
            fac_id = f"fac_{make_id([norm])}"
            facility = {
                "id": fac_id,
                "name": name.strip(),
                "aliases": [],
                "address": addr_norm,
                "lga": lga,
                "zone": zone or LGA_ZONE_MAP.get(lga.upper(), "") if lga else "",
                "facilityType": facility_type,
                "activityTypes": [activity_type] if activity_type else [],
                "fileNumber": file_number,
                "contactPerson": contact_person,
                "phone": phone,
                "email": email,
                "totalVisits": 0,
                "lastVisitDate": "",
                "outstandingFines": 0,
                "totalFinesIssued": 0,
                "totalFinesPaid": 0,
                "status": "Active",
            }
            self.facilities[norm] = facility
            return fac_id

    def get_id(self, name):
        """Look up a facility ID by name."""
        norm = normalize_name(name)
        match_key = self._match_key(name)

        if norm in self.facilities:
            return self.facilities[norm]["id"]
        for existing_norm, facility in self.facilities.items():
            if self._match_key(facility["name"]) == match_key and match_key:
                return facility["id"]
        return None

    def get_all(self):
        """Return all facilities as a list."""
        return list(self.facilities.values())


# ─── Extractors ──────────────────────────────────────────────────────────────

def extract_routine_surveillance(registry):
    """Extract from PMS-LAGOS ROUTINE SURVEILLANCE (2016–2025)."""
    print("\n📋 Extracting Routine Surveillance...")
    wb = read_workbook("routine_surveillance")
    if not wb:
        return []

    inspections = []
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        year_match = re.search(r'20\d{2}', sheet_name)
        year = int(year_match.group()) if year_match else 0

        rows = list(ws.iter_rows(values_only=True))
        # Find header row
        header_row = 0
        for i, row in enumerate(rows):
            if row and any(str(c).upper().startswith('S/N') for c in row if c):
                header_row = i
                break

        for row in rows[header_row + 1:]:
            if not row or not row[1]:  # Skip empty rows (NAME column)
                continue

            name = cell_str(row[1])
            if not name or name.upper() == 'NAME':
                continue

            address = cell_str(row[2]) if len(row) > 2 else ""
            date_val = cell_str(row[3]) if len(row) > 3 else ""
            observation = cell_str(row[4]) if len(row) > 4 else ""
            action = cell_str(row[5]) if len(row) > 5 else ""
            recommendation = cell_str(row[6]) if len(row) > 6 else ""
            admin_charge = cell_str(row[7]) if len(row) > 7 else ""
            warning_letter = cell_str(row[8]) if len(row) > 8 else ""
            remark = cell_str(row[9]) if len(row) > 9 else ""

            fac_id = registry.register(name, address=address, activity_type="Routine Surveillance")

            inspections.append({
                "facilityId": fac_id,
                "facilityName": name.strip(),
                "activityType": "Routine Surveillance",
                "inspectionDate": parse_date(date_val),
                "year": year,
                "observation": observation,
                "actionTaken": action,
                "recommendation": recommendation,
                "adminCharge": admin_charge,
                "warningLetter": warning_letter,
                "remark": remark,
                "status": remark.upper() if remark else "OPEN",
                "source": f"routine_surveillance/{sheet_name}"
            })
            total += 1

    wb.close()
    print(f"  ✅ {total} inspection records from {len(wb.sheetnames)} yearly sheets")
    return inspections


def extract_gdp_inspected(registry):
    """Extract from GDP UPDATED INSPECTED FACILITIES."""
    print("\n🏭 Extracting GDP Inspected Facilities...")
    wb = read_workbook("gdp_inspected")
    if not wb:
        return []

    inspections = []
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        year_match = re.search(r'20\d{2}', sheet_name)
        year = int(year_match.group()) if year_match else 0

        rows = list(ws.iter_rows(values_only=True))
        # Find header row (contains 'S/N')
        header_row = 0
        for i, row in enumerate(rows):
            if row and any(str(c).upper().startswith('S/N') for c in row if c):
                header_row = i
                break

        for row in rows[header_row + 1:]:
            if not row or len(row) < 6:
                continue

            name = cell_str(row[1])
            if not name or name.upper() in ('NAME OF MANUFACTURER', ''):
                continue

            address = cell_str(row[2])
            contact = cell_str(row[3])
            inspection_type = cell_str(row[4])
            date_val = cell_str(row[5])
            findings = cell_str(row[6]) if len(row) > 6 else ""
            capa_issued = cell_str(row[7]) if len(row) > 7 else ""
            capa_submitted = cell_str(row[8]) if len(row) > 8 else ""
            conclusion = cell_str(row[9]) if len(row) > 9 else ""
            remarks = cell_str(row[10]) if len(row) > 10 else ""
            expected_date = cell_str(row[11]) if len(row) > 11 else ""

            # Try to get SharePoint link — might be in col 12 or 13
            company_file_link = ""
            for ci in range(12, min(len(row), 15)):
                v = cell_str(row[ci])
                if v and v.startswith("http"):
                    company_file_link = v
                    break

            fac_id = registry.register(
                name, address=address, contact_person=contact,
                activity_type="GSDP", facility_type="distributor"
            )

            inspections.append({
                "facilityId": fac_id,
                "facilityName": name.strip(),
                "activityType": "GSDP",
                "subActivity": inspection_type or "GDP",
                "inspectionDate": parse_date(date_val),
                "year": year,
                "riskFinding": findings,
                "capaIssuedDate": parse_date(capa_issued),
                "capaSubmitted": capa_submitted,
                "capaConclusion": conclusion,
                "remark": remarks,
                "expectedNextInspection": parse_date(expected_date),
                "companyFileLink": company_file_link,
                "status": "CLOSED" if "submitted" in conclusion.lower() or "verified" in conclusion.lower() else "OPEN",
                "source": f"gdp_inspected/{sheet_name}"
            })
            total += 1

    wb.close()
    print(f"  ✅ {total} GDP inspection records from {len(wb.sheetnames)} sheets")
    return inspections


def extract_gdp_uninspected(registry):
    """Extract from CURRENT UNINSPECTED GDP UPDATE."""
    print("\n🔍 Extracting GDP Uninspected Facilities...")
    wb = read_workbook("gdp_uninspected")
    if not wb:
        return []

    inspections = []
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        year_match = re.search(r'20\d{2}', sheet_name)
        year = int(year_match.group()) if year_match else 0

        rows = list(ws.iter_rows(values_only=True))
        header_row = 0
        for i, row in enumerate(rows):
            if row and any(str(c).upper().startswith('S/N') for c in row if c):
                header_row = i
                break

        for row in rows[header_row + 1:]:
            if not row or len(row) < 4:
                continue
            name = cell_str(row[1])
            if not name or name.upper() in ('NAME OF MANUFACTURER', ''):
                continue

            address = cell_str(row[2])
            contact = cell_str(row[3])
            product_lines = cell_str(row[4]) if len(row) > 4 else ""
            insp_type = cell_str(row[5]) if len(row) > 5 else "GDP"
            date_val = cell_str(row[6]) if len(row) > 6 else ""
            status_or_zone = cell_str(row[7]) if len(row) > 7 else ""
            capa_or_remarks = cell_str(row[8]) if len(row) > 8 else ""
            zone = cell_str(row[9]) if len(row) > 9 else ""
            remarks = cell_str(row[10]) if len(row) > 10 else ""

            # Column layout differs by year — handle both layouts
            if year <= 2023:
                status_val = status_or_zone
                capa_status = capa_or_remarks
            else:
                status_val = status_or_zone
                capa_status = capa_or_remarks

            fac_id = registry.register(
                name, address=address, contact_person=contact,
                activity_type="GSDP", facility_type="distributor"
            )

            inspections.append({
                "facilityId": fac_id,
                "facilityName": name.strip(),
                "activityType": "GSDP",
                "subActivity": insp_type or "GDP",
                "inspectionDate": parse_date(date_val),
                "year": year,
                "productLines": product_lines,
                "facilityStatus": status_val,
                "capaStatus": capa_status,
                "remark": remarks,
                "status": "UNINSPECTED",
                "source": f"gdp_uninspected/{sheet_name}"
            })
            total += 1

    wb.close()
    print(f"  ✅ {total} uninspected GDP facility records")
    return inspections


def extract_glsi(registry):
    """Extract GLSI data from the three zone files + not-located."""
    print("\n🏪 Extracting GLSI Data...")
    inspections = []
    total = 0

    zone_files = {
        "glsi_central": "Lagos Central",
        "glsi_east": "Lagos East",
        "glsi_west": "Lagos West",
    }

    for key, zone_name in zone_files.items():
        wb = read_workbook(key)
        if not wb:
            continue

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lga = sheet_name.upper().strip()

            rows = list(ws.iter_rows(values_only=True))
            if not rows or ws.max_row <= 1:
                continue

            header_row = 0
            for i, row in enumerate(rows):
                if row and any(str(c).upper().startswith('S/N') for c in row if c):
                    header_row = i
                    break

            for row in rows[header_row + 1:]:
                if not row or len(row) < 3:
                    continue
                # Name is usually at index 1
                name = cell_str(row[1])
                if not name or name.upper() in ('NAME', ''):
                    continue

                address = cell_str(row[2]) if len(row) > 2 else ""
                date_val = cell_str(row[3]) if len(row) > 3 else ""
                observation = cell_str(row[4]) if len(row) > 4 else ""
                action = cell_str(row[5]) if len(row) > 5 else ""
                recommendation = cell_str(row[6]) if len(row) > 6 else ""

                fac_id = registry.register(
                    name, address=address, activity_type="GLSI",
                    zone=zone_name, lga=lga, facility_type="supermarket"
                )

                inspections.append({
                    "facilityId": fac_id,
                    "facilityName": name.strip(),
                    "activityType": "GLSI",
                    "glsiZone": zone_name,
                    "lga": lga,
                    "inspectionDate": parse_date(date_val),
                    "year": 0,  # Will try to infer from date
                    "observation": observation,
                    "actionTaken": action,
                    "recommendation": recommendation,
                    "status": "CLOSED",
                    "source": f"{key}/{sheet_name}"
                })
                total += 1
        wb.close()

    # GLSI Not Located
    wb = read_workbook("glsi_not_located")
    if wb:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            header_row = 0
            for i, row in enumerate(rows):
                if row and any(str(c).upper().startswith('S/N') for c in row if c):
                    header_row = i
                    break

            for row in rows[header_row + 1:]:
                if not row or len(row) < 3:
                    continue
                name = cell_str(row[1])
                if not name:
                    continue
                address = cell_str(row[2]) if len(row) > 2 else ""
                date_val = cell_str(row[3]) if len(row) > 3 else ""
                observation = cell_str(row[4]) if len(row) > 4 else ""

                fac_id = registry.register(
                    name, address=address, activity_type="GLSI"
                )
                # Mark facility as not located
                norm = normalize_name(name)
                if norm in registry.facilities:
                    registry.facilities[norm]["status"] = "Not Located"

                inspections.append({
                    "facilityId": fac_id,
                    "facilityName": name.strip(),
                    "activityType": "GLSI",
                    "inspectionDate": parse_date(date_val),
                    "observation": observation or "Facility not located at stated address",
                    "status": "NOT_LOCATED",
                    "source": f"glsi_not_located/{sheet_name}"
                })
                total += 1
        wb.close()

    print(f"  ✅ {total} GLSI records across all zones")
    return inspections


def extract_sanctions(registry):
    """Extract sanctions from Admin Fees, Revenue, Defaulters, GLSI Defaulters."""
    print("\n💰 Extracting Sanctions & Financial Records...")
    sanctions = []
    total = 0

    # ── Admin Fees 2021-2025 ──
    wb = read_workbook("admin_fees")
    if wb:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            year_match = re.search(r'20\d{2}', sheet_name)
            year = int(year_match.group()) if year_match else 0

            rows = list(ws.iter_rows(values_only=True))
            header_row = 0
            for i, row in enumerate(rows):
                if row and any('NAMES OF COMPANY' in str(c).upper() for c in row if c):
                    header_row = i
                    break
                if row and any('NATURE OF OFFENCE' in str(c).upper() for c in row if c):
                    header_row = i
                    break

            for row in rows[header_row + 1:]:
                if not row or not row[0]:
                    continue
                name = cell_str(row[0])
                if not name or 'NAMES OF COMPANY' in name.upper() or 'COMPANIES THAT' in name.upper():
                    continue

                offence = cell_str(row[1]) if len(row) > 1 else ""
                amount = parse_amount(row[2]) if len(row) > 2 else 0
                status = cell_str(row[4]) if len(row) > 4 else ""
                date_issued = cell_str(row[5]) if len(row) > 5 else ""
                address = cell_str(row[6]) if len(row) > 6 else ""
                contact = cell_str(row[7]) if len(row) > 7 else ""
                phone = cell_str(row[8]) if len(row) > 8 else ""

                fac_id = registry.register(
                    name, address=address, contact_person=contact, phone=phone
                )

                sanctions.append({
                    "facilityId": fac_id,
                    "facilityName": name.strip(),
                    "sanctionType": "Admin Fee",
                    "offence": offence,
                    "amount": amount,
                    "paymentStatus": status.upper() if status else "PENDING",
                    "issuedDate": parse_date(date_issued),
                    "year": year,
                    "contactPerson": contact,
                    "phone": phone,
                    "companyAddress": address,
                    "source": f"admin_fees/{sheet_name}"
                })
                total += 1
        wb.close()

    # ── Revenue Sheet 2025 ──
    wb = read_workbook("revenue")
    if wb:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            header_row = 0
            for i, row in enumerate(rows):
                if row and any('NAMES OF COMPANY' in str(c).upper() for c in row if c):
                    header_row = i
                    break
                if row and any('NATURE OF OFFENCE' in str(c).upper() for c in row if c):
                    header_row = i
                    break

            for row in rows[header_row + 1:]:
                if not row or not row[0]:
                    continue
                name = cell_str(row[0])
                if not name or 'NAMES OF COMPANY' in name.upper() or 'COMPANIES THAT' in name.upper():
                    continue

                offence = cell_str(row[1]) if len(row) > 1 else ""
                amount = parse_amount(row[2]) if len(row) > 2 else 0
                status = cell_str(row[3]) if len(row) > 3 else ""

                # Determine month/year — layout varies per sheet
                month = ""
                year_val = 2025
                if len(row) > 4:
                    month = cell_str(row[4])
                if len(row) > 5:
                    try:
                        y = int(row[5])
                        if 2020 <= y <= 2030:
                            year_val = y
                    except (ValueError, TypeError):
                        pass

                fac_id = registry.register(name)

                sanctions.append({
                    "facilityId": fac_id,
                    "facilityName": name.strip(),
                    "sanctionType": "Admin Fee",
                    "offence": offence,
                    "amount": amount,
                    "paymentStatus": status.upper() if status else "PENDING",
                    "month": month,
                    "year": year_val,
                    "source": f"revenue/{sheet_name}"
                })
                total += 1
        wb.close()

    # ── Defaulters 2021-2025 ──
    wb = read_workbook("defaulters")
    if wb:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            year_match = re.search(r'20\d{2}', sheet_name)
            year = int(year_match.group()) if year_match else 0

            rows = list(ws.iter_rows(values_only=True))
            header_row = 0
            for i, row in enumerate(rows):
                if row and any('NAMES OF COMPANY' in str(c).upper() for c in row if c):
                    header_row = i
                    break
                if row and any('NATURE OF OFFENCE' in str(c).upper() for c in row if c):
                    header_row = i
                    break

            for row in rows[header_row + 1:]:
                if not row or len(row) < 2:
                    continue
                name = cell_str(row[0])
                if not name or 'NAMES' in name.upper() or 'DEFAULTER' in name.upper():
                    continue

                offence = cell_str(row[1]) if len(row) > 1 else ""
                amount = parse_amount(row[2]) if len(row) > 2 else 0
                contact = cell_str(row[3]) if len(row) > 3 else ""
                address = cell_str(row[4]) if len(row) > 4 else ""
                phone = cell_str(row[5]) if len(row) > 5 else ""
                status = cell_str(row[6]) if len(row) > 6 else "PENDING"

                fac_id = registry.register(
                    name, address=address, contact_person=contact, phone=phone
                )

                sanctions.append({
                    "facilityId": fac_id,
                    "facilityName": name.strip(),
                    "sanctionType": "Admin Fee",
                    "offence": offence,
                    "amount": amount,
                    "paymentStatus": "PENDING",
                    "year": year,
                    "contactPerson": contact,
                    "phone": phone,
                    "companyAddress": address,
                    "defaulter": True,
                    "source": f"defaulters/{sheet_name}"
                })
                total += 1
        wb.close()

    # ── GLSI Defaulters ──
    wb = read_workbook("glsi_defaulters")
    if wb:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            header_row = 0
            for i, row in enumerate(rows):
                if row and any('NAMES OF COMPANY' in str(c).upper() for c in row if c):
                    header_row = i
                    break

            for row in rows[header_row + 1:]:
                if not row or not row[0]:
                    continue
                name = cell_str(row[0])
                if not name or 'NAMES' in name.upper():
                    continue

                offence = cell_str(row[1]) if len(row) > 1 else ""
                amount = parse_amount(row[2]) if len(row) > 2 else 0
                status = cell_str(row[3]) if len(row) > 3 else ""
                year_val = cell_str(row[4]) if len(row) > 4 else ""

                fac_id = registry.register(name, activity_type="GLSI")

                sanctions.append({
                    "facilityId": fac_id,
                    "facilityName": name.strip(),
                    "sanctionType": "Admin Fee (GLSI)",
                    "offence": offence,
                    "amount": amount,
                    "paymentStatus": status.upper() if status else "PENDING",
                    "year": int(year_val) if year_val.isdigit() else 0,
                    "source": f"glsi_defaulters/{sheet_name}"
                })
                total += 1
        wb.close()

    print(f"  ✅ {total} sanction/financial records")
    return sanctions


def extract_complaints(registry):
    """Extract from CONSUMER COMPLAINTS (2018–2027)."""
    print("\n📞 Extracting Consumer Complaints...")
    wb = read_workbook("complaints")
    if not wb:
        return []

    complaints = []
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        year_match = re.search(r'20\d{2}', sheet_name)
        year = int(year_match.group()) if year_match else 0

        rows = list(ws.iter_rows(values_only=True))
        if not rows or ws.max_row <= 1:
            continue

        # Different formats across years — detect by headers
        header_row = 0
        for i, row in enumerate(rows):
            if row and any(str(c).upper().strip() in ('S/N', 'MONTH/ YEAR', 'DATE') for c in row if c):
                header_row = i
                break

        headers = [cell_str(c).upper() for c in rows[header_row]] if header_row < len(rows) else []

        for row in rows[header_row + 1:]:
            if not row or all(c is None for c in row):
                continue

            # Try to extract based on detected format
            if 'PRODUCT COMPLAINT' in headers or 'S/N' in headers:
                # 2018 format: S/N, PRODUCT COMPLAINT, OUTLET VISITED, OBSERVATION, TYPE, OUTCOME
                product = cell_str(row[1]) if len(row) > 1 else ""
                outlet = cell_str(row[2]) if len(row) > 2 else ""
                observation = cell_str(row[3]) if len(row) > 3 else ""
                prod_type = cell_str(row[4]) if len(row) > 4 else ""
                outcome = cell_str(row[5]) if len(row) > 5 else ""
                date_val = ""
                complaint = ""
                action = ""
            else:
                # 2019+ format: Month/year, Date, Type/Source, Product, NRN, Complaint, Observations, Action, ...
                date_val = cell_str(row[1]) if len(row) > 1 else ""
                prod_type = cell_str(row[2]) if len(row) > 2 else ""
                product = cell_str(row[3]) if len(row) > 3 else ""
                complaint = cell_str(row[5]) if len(row) > 5 else ""
                observation = cell_str(row[6]) if len(row) > 6 else ""
                action = cell_str(row[7]) if len(row) > 7 else ""
                outlet = ""
                outcome = cell_str(row[9]) if len(row) > 9 else ""

            if not product and not observation:
                continue

            # Try to link to a facility if outlet name is mentioned
            fac_id = None
            if outlet:
                fac_id = registry.get_id(outlet)
                if not fac_id:
                    fac_id = registry.register(outlet, activity_type="Consumer Complaint")

            complaints.append({
                "facilityId": fac_id,
                "product": product,
                "productType": prod_type,
                "complaint": complaint,
                "outletVisited": outlet,
                "observation": observation,
                "actionTaken": action,
                "outcome": outcome.upper() if outcome else "",
                "year": year,
                "dateReceived": parse_date(date_val),
                "source": f"complaints/{sheet_name}"
            })
            total += 1

    wb.close()
    print(f"  ✅ {total} consumer complaint records")
    return complaints


def extract_documents(registry):
    """Extract from DOCUMENTS RECEIVED AND DISPATCHED IN 2025."""
    print("\n📄 Extracting Document Correspondence...")
    wb = read_workbook("documents")
    if not wb:
        return []

    documents = []
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        direction = "RECEIVED" if "RECEIVED" in sheet_name.upper() else "DISPATCHED"

        rows = list(ws.iter_rows(values_only=True))
        header_row = 0
        for i, row in enumerate(rows):
            if row and any('DATE' in str(c).upper() for c in row if c):
                header_row = i
                break

        for row in rows[header_row + 1:]:
            if not row or not row[0]:
                continue

            date_val = cell_str(row[0])
            if not date_val or 'DATE' in date_val.upper():
                continue

            if direction == "RECEIVED":
                sender = cell_str(row[1]) if len(row) > 1 else ""
                subject = cell_str(row[2]) if len(row) > 2 else ""
                addressed_to = cell_str(row[3]) if len(row) > 3 else ""
                receiver = cell_str(row[4]) if len(row) > 4 else ""
                remark = cell_str(row[5]) if len(row) > 5 else ""
                minuted_to = cell_str(row[6]) if len(row) > 6 else ""

                documents.append({
                    "direction": direction,
                    "date": parse_date(date_val),
                    "sender": sender,
                    "subject": subject,
                    "addressedTo": addressed_to,
                    "receiver": receiver,
                    "remark": remark,
                    "minutedTo": minuted_to,
                    "year": 2025,
                    "source": f"documents/{sheet_name}"
                })
            else:
                dispatcher = cell_str(row[1]) if len(row) > 1 else ""
                subject = cell_str(row[2]) if len(row) > 2 else ""
                sender_name = cell_str(row[3]) if len(row) > 3 else ""
                remark = cell_str(row[4]) if len(row) > 4 else ""

                documents.append({
                    "direction": direction,
                    "date": parse_date(date_val),
                    "sender": sender_name,
                    "dispatcher": dispatcher,
                    "subject": subject,
                    "remark": remark,
                    "year": 2025,
                    "source": f"documents/{sheet_name}"
                })
            total += 1

    wb.close()
    print(f"  ✅ {total} document records")
    return documents


def extract_file_registry(registry):
    """Extract from PMS LAGOS OFFICE FILE."""
    print("\n📁 Extracting File Registry...")
    wb = read_workbook("office_files")
    if not wb:
        return []

    file_records = []
    total = 0

    category_map = {
        "GLSI FILES": "GLSI",
        "PMS FILES": "PMS",
        "GDP FILES": "GDP",
        "GENERAL FILES": "GENERAL",
        "PVG FILES": "PVG",
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        category = category_map.get(sheet_name, "OTHER")

        rows = list(ws.iter_rows(values_only=True))
        header_row = 0
        for i, row in enumerate(rows):
            if row and any('FILE NAME' in str(c).upper() for c in row if c):
                header_row = i
                break

        for row in rows[header_row + 1:]:
            if not row or not row[0]:
                continue
            name = cell_str(row[0])
            if not name or 'FILE NAME' in name.upper() or 'PMS LAGOS' in name.upper():
                continue

            file_number = cell_str(row[1]) if len(row) > 1 else ""
            volume = cell_str(row[2]) if len(row) > 2 else ""

            # Register the facility with its file number
            fac_id = registry.register(name, file_number=file_number)

            file_records.append({
                "facilityId": fac_id,
                "fileName": name.strip(),
                "fileNumber": file_number,
                "fileCategory": category,
                "volumeNumber": volume,
                "source": f"office_files/{sheet_name}"
            })
            total += 1

    wb.close()
    print(f"  ✅ {total} file registry records across {len(wb.sheetnames)} categories")
    return file_records


# ─── Aggregation ─────────────────────────────────────────────────────────────

def aggregate_facility_stats(registry, inspections, sanctions):
    """Update facility master records with aggregated statistics."""
    print("\n📊 Aggregating facility statistics...")

    # Count inspections per facility
    visit_counts = defaultdict(int)
    last_dates = {}
    for insp in inspections:
        fid = insp.get("facilityId")
        if not fid:
            continue
        visit_counts[fid] += 1
        d = insp.get("inspectionDate", "")
        if d and (fid not in last_dates or d > last_dates[fid]):
            last_dates[fid] = d

    # Aggregate sanctions
    fines_issued = defaultdict(int)
    fines_paid = defaultdict(int)
    outstanding = defaultdict(int)
    for sn in sanctions:
        fid = sn.get("facilityId")
        if not fid:
            continue
        amt = sn.get("amount", 0)
        fines_issued[fid] += amt
        if sn.get("paymentStatus") == "PAID":
            fines_paid[fid] += amt
        else:
            outstanding[fid] += amt

    # Update facilities
    for norm, fac in registry.facilities.items():
        fid = fac["id"]
        fac["totalVisits"] = visit_counts.get(fid, 0)
        fac["lastVisitDate"] = last_dates.get(fid, "")
        fac["totalFinesIssued"] = fines_issued.get(fid, 0)
        fac["totalFinesPaid"] = fines_paid.get(fid, 0)
        fac["outstandingFines"] = outstanding.get(fid, 0)

    print(f"  ✅ Updated stats for {len(registry.facilities)} facilities")


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("PMS LAGOS — Unified Facility Database ETL")
    print("=" * 60)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    registry = FacilityRegistry()

    # 1. Extract all data
    inspections = []
    inspections += extract_routine_surveillance(registry)
    inspections += extract_gdp_inspected(registry)
    inspections += extract_gdp_uninspected(registry)
    inspections += extract_glsi(registry)

    sanctions = extract_sanctions(registry)
    complaints = extract_complaints(registry)
    documents = extract_documents(registry)
    file_registry = extract_file_registry(registry)

    # 2. Aggregate statistics
    aggregate_facility_stats(registry, inspections, sanctions)

    # 3. Write output files
    facilities = registry.get_all()

    outputs = {
        "master_facilities.json": facilities,
        "inspections.json": inspections,
        "sanctions.json": sanctions,
        "complaints.json": complaints,
        "documents.json": documents,
        "file_registry.json": file_registry,
        "dedup_report.json": registry.merge_log,
    }

    print("\n" + "=" * 60)
    print("📦 Writing output files...")
    for filename, data in outputs.items():
        path = os.path.join(OUTPUT_DIR, filename)
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
        print(f"  📄 {filename}: {len(data)} records")

    print("\n" + "=" * 60)
    print("✅ ETL COMPLETE — Summary:")
    print(f"  🏢 Unique facilities:     {len(facilities)}")
    print(f"  📋 Inspection records:    {len(inspections)}")
    print(f"  💰 Sanction records:      {len(sanctions)}")
    print(f"  📞 Complaint records:     {len(complaints)}")
    print(f"  📄 Document records:      {len(documents)}")
    print(f"  📁 File registry records: {len(file_registry)}")
    print(f"  🔀 Merges performed:      {len(registry.merge_log)}")
    print(f"\n  Output directory: {OUTPUT_DIR}")
    print("=" * 60)


if __name__ == "__main__":
    main()
